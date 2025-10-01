#!/usr/bin/env python3

import os
import sys
import json
import csv
import argparse
from datetime import datetime
from typing import Dict, Any, List, Optional

# Import URL parser with flexible execution contexts
try:
    from src.utils.paths import parse_tiktok_url
except ModuleNotFoundError:
    THIS_DIR = os.path.dirname(os.path.abspath(__file__))
    PROJECT_ROOT = os.path.abspath(os.path.join(THIS_DIR, os.pardir, os.pardir))
    if PROJECT_ROOT not in sys.path:
        sys.path.insert(0, PROJECT_ROOT)
    try:
        from src.utils.paths import parse_tiktok_url
    except ModuleNotFoundError:
        from ..utils.paths import parse_tiktok_url  # type: ignore


def _to_iso(ts: Any) -> Optional[str]:
    try:
        return datetime.utcfromtimestamp(int(ts)).isoformat() + "Z"
    except Exception:
        return None


def _normalize_comment(obj: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    # Try a variety of common key names
    cid = obj.get("cid") or obj.get("id") or obj.get("comment_id")
    text = obj.get("text") or obj.get("content") or obj.get("comment")
    # author can be nested
    author = (
        (obj.get("user") or {}).get("unique_id")
        or (obj.get("user") or {}).get("username")
        or (obj.get("author") or {}).get("unique_id")
        or obj.get("username")
        or obj.get("author")
    )
    like_count = obj.get("digg_count") or obj.get("like_count") or obj.get("likes")
    create_time = obj.get("create_time") or obj.get("timestamp")

    if not text:
        return None
    if not cid:
        cid = f"{author}-{hash(text)}"
    return {
        "id": cid,
        "author": author,
        "text": text,
        "like_count": like_count,
        "create_time": create_time,
    }


def _load_comments(payload: Any) -> List[Dict[str, Any]]:
    # Accept various shapes: {comments: [...]}, {data: {comments: [...]}}, or list
    if isinstance(payload, list):
        items = payload
    elif isinstance(payload, dict):
        if isinstance(payload.get("comments"), list):
            items = payload["comments"]
        elif isinstance(payload.get("data"), dict) and isinstance(payload["data"].get("comments"), list):
            items = payload["data"]["comments"]
        else:
            # Try common alternative container keys
            for key in ("items", "results", "records"):
                if isinstance(payload.get(key), list):
                    items = payload[key]
                    break
            else:
                items = []
    else:
        items = []

    def _flatten_thirdparty(it: Dict[str, Any], out: List[Dict[str, Any]], seen: set, parent: Optional[str] = None):
        norm = _normalize_comment(it)
        if norm:
            if parent:
                norm["parent_id"] = parent
            cid = norm.get("id")
            if cid not in seen:
                seen.add(cid)
                out.append(norm)
        # Recurse into replies if present
        reps = it.get("replies")
        if isinstance(reps, list):
            pid = it.get("comment_id") or it.get("cid") or it.get("id") or (norm and norm.get("id"))
            for r in reps:
                if isinstance(r, dict):
                    _flatten_thirdparty(r, out, seen, parent=pid)

    out: List[Dict[str, Any]] = []
    seen = set()
    for it in items:
        if not isinstance(it, dict):
            continue
        if ("comment_id" in it) or ("replies" in it):
            _flatten_thirdparty(it, out, seen)
        else:
            norm = _normalize_comment(it)
            if not norm:
                continue
            cid = norm.get("id")
            if cid in seen:
                continue
            seen.add(cid)
            out.append(norm)
    return out


def _load_from_csv(input_path: str) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with open(input_path, "r", encoding="utf-8") as f:
        r = csv.reader(f)
        header_found = False
        headers: List[str] = []
        for row in r:
            if not header_found:
                # Detect the header row by its first column prefix
                if row and row[0].strip().lower().startswith("comment number"):
                    headers = [c.strip() for c in row]
                    header_found = True
                continue
            if not row:
                continue
            # Map by known header names
            rec = dict(zip(headers, row))
            rows.append(rec)
    return rows


def _load_from_xlsx(input_path: str) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError("openpyxl is required to read .xlsx files. Install it first.") from e
    wb = load_workbook(input_path, read_only=True)
    ws = wb.active
    rows: List[Dict[str, Any]] = []
    header_found = False
    headers: List[str] = []
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row and isinstance(row[0], str) and row[0].strip().lower().startswith("comment number"):
                headers = [str(c) if c is not None else "" for c in row]
                header_found = True
            continue
        if not row:
            continue
        rec = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        rows.append(rec)
    return rows


def _normalize_tpcs_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Normalize rows from TikTokCommentScraper CSV/XLSX to our schema."""
    out: List[Dict[str, Any]] = []
    seen = set()
    for rec in rows:
        author = rec.get("User @") or rec.get("User @ ") or rec.get("User")
        text = rec.get("Comment Text")
        like_count = rec.get("Likes")
        created = rec.get("Time")
        cid = rec.get("Comment Number (ID)")
        if text is None or text == "":
            continue
        if not cid:
            cid = f"{author}-{hash(text)}"
        if cid in seen:
            continue
        seen.add(cid)
        out.append({
            "id": cid,
            "author": author,
            "text": text,
            "like_count": like_count,
            "create_time": created,
        })
    return out


def ingest(input_path: str, url: str, out_dir: str, limit: int = 0) -> str:
    uploader, vid = parse_tiktok_url(url)
    base = f"{uploader}_video_{vid}" if uploader and vid else "tiktok_video"
    os.makedirs(out_dir, exist_ok=True)

    ext = os.path.splitext(input_path)[1].lower()
    if ext == ".json":
        with open(input_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        rows = _load_comments(data)
    elif ext == ".csv":
        trows = _load_from_csv(input_path)
        rows = _normalize_tpcs_rows(trows)
    elif ext in (".xlsx", ".xlsm"):
        trows = _load_from_xlsx(input_path)
        rows = _normalize_tpcs_rows(trows)
    else:
        raise ValueError(f"Unsupported input format: {ext}")

    summary_path = os.path.join(out_dir, f"{base}_comments.json")
    jsonl_path = os.path.join(out_dir, f"{base}_comments.jsonl")

    # Optional limit
    if limit and isinstance(rows, list):
        rows = rows[:limit]

    # Write summary JSON
    payload = {
        "video_url": url,
        "uploader": uploader,
        "video_id": vid,
        "count": len(rows),
        "comments": rows,
    }
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    # Write JSONL
    now = datetime.utcnow().isoformat() + "Z"
    with open(jsonl_path, "w", encoding="utf-8") as f:
        for r in rows:
            rec = dict(r)
            rec["video_id"] = vid
            rec["uploader"] = uploader
            rec["source_url"] = url
            rec["collected_at"] = now
            if rec.get("create_time"):
                rec["create_time_iso"] = _to_iso(rec["create_time"])
            rec["parser_version"] = "tp1"
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    return summary_path


def main():
    parser = argparse.ArgumentParser(description="Ingest third-party TikTok comments JSON and normalize to project schema")
    parser.add_argument("input", help="Path to third-party comments JSON file")
    parser.add_argument("--url", required=True, help="TikTok video URL (used for naming and metadata)")
    parser.add_argument("--out", required=True, help="Output directory (e.g., runs/<...>)")
    parser.add_argument("--limit", type=int, default=0, help="Limit number of comments (0 = no limit)")
    args = parser.parse_args()

    out = ingest(args.input, args.url, args.out, limit=args.limit)
    print(f"Ingested comments -> {out}")


if __name__ == "__main__":
    main()
